# -*- coding: utf-8 -*-
#

import os
import re
from openpyxl import load_workbook

def iterate_all_xlsx(inpath):
    if os.path.isfile(inpath):
        yield inpath
    else:
        for f in os.listdir(inpath):
            if os.path.splitext(f)[1] == '.xlsx':
                yield os.path.join(inpath, f)

def retrieve_contents(inpath, outfile):
    regex = re.compile('\r\n|\n')
    with open(outfile, 'w') as of:
        for infile in iterate_all_xlsx(inpath):
            #print(infile)
            wb = load_workbook(infile)
            ws = wb.active
            cells = ws['C']
            #print(len(cells) - 1)
            for cell in cells:
                if cell.row < 2:
                    continue
                of.write(regex.sub(' ', str(cell.value)) + '\n')

def machine_filter(infile, outfile, kwfile):
    from functools import reduce
    kwdict = {}
    with open(kwfile, 'r') as kf:
        for w in kf:
            print(w[2:], w[:1])
            kwdict[w[2:].strip()] = w[:1]
    print(kwdict)
    regstr = reduce(lambda r, w : r + '|' + w, kwdict.keys())
    print(regstr)
    regex = re.compile(regstr)

    GOOD_CONTENT, BAD_CONTENT, SKIP_CONTENT = '1', '2', '3'
    count = 0
    with open(outfile, 'w') as rf:
        with open(infile, 'r') as df:
            for i, line in enumerate(df):
                #if regex.match(line):
                m = regex.search(line)
                if m:
                    rf.write(kwdict[m.group(0)] + '\n')
                    print(kwdict[m.group(0)], m.group(0))
                    count += 1
                else:
                    #pass
                    rf.write(SKIP_CONTENT + '\n')
                    #print('3')
    print('Matched %d lines.' % count) 

