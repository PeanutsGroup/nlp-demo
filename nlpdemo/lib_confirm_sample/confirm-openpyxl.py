#!/usr/bin/env python
# -*- coding: utf-8 -*-
#

from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("/tmp/sample.xlsx")

from openpyxl import load_workbook
wb2 = load_workbook('/tmp/sample.xlsx')
print(wb2.get_sheet_names())

